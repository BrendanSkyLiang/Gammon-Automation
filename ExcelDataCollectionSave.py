# -*- coding: utf-8 -*-
"""
Created on Wed Jul  7 09:17:36 2021

@author: brendanlia
"""

import numpy
import xlsxwriter
import pandas
import openpyxl
from openpyxl import load_workbook
import time
import win32com.client
import xlrd
from win32com import client

'--------------------------------------------------------------------------------------------------------------------------------------------------'
#Def Functions

def refreshExcel():
    xlapp = win32com.client.DispatchEx("Excel.Application")
    # The File location will need to be edited for different file locations
    wb = xlapp.Workbooks.Open(r"C:\Users\brendanlia\Desktop\URGENT\BoltCalc3.xlsx") 
    wb.RefreshAll()
    xlapp.CalculateUntilAsyncQueriesDone()
    wb.Save()
    xlapp.Quit()
    
    
def refreshExcel1(filePath):
    xlapp = win32com.client.DispatchEx("Excel.Application")
    # The File location will need to be edited for different file locations
    wb = xlapp.Workbooks.Open(filePath) 
    wb.RefreshAll()
    xlapp.CalculateUntilAsyncQueriesDone()
    wb.Save()
    xlapp.Quit()
'----------------------------------------------------------------------------------------------------------------------------------------------------'
    
def getMaxUR(ForceChoice, BoltCalcStr, ForceCombinedStr, ForceStr):
    # ForceChoice = selction of MaxFx, MaxFy, ... , MinFy
    # BoltCalcStr = file name of the bolt calc spreadsheet, input as a string e.g. "BoltCalc.xlsx"
    # ForceCombinedStr = the file name of the combined forces xlsx doc, input as string, e.g. "CombinedMinMax.xlsx"
    # ForceStr = the Force Choice in string format, input as string, e.g. "MaxFx"
    
    TUR = []
    BUR = []
    SUR = []
    
    for a in range(len(ForceChoice.columns)): #Input new Force Values
        
        Group = ForceChoice[a][11]
        size = len(Group)-2
        cut = Group[2:][:size - 2]
        sheetCalcStr = cut
        
        workbook = load_workbook(filename = BoltCalcStr)
        sheet = workbook[sheetCalcStr]
        if ForceChoice[a][7] > 0: #Compression
            sheet["F6"] = 0
            sheet["F7"] = abs(ForceChoice[a][7]) 
        elif ForceChoice[a][7] < 0: #Tension
            sheet["F6"] = abs(ForceChoice[a][7])
            sheet["F7"] = 0
        else:     
            sheet["F6"] = 0
            sheet["F7"] = 0
        
        sheet["F8"] = ForceChoice[a][5]
        sheet["F9"] = ForceChoice[a][6] 
        sheet["F10"] = abs(ForceChoice[a][9])
        sheet["F11"] = abs(ForceChoice[a][8])
        
        sheet['L1'] = ForceChoice[a][3] # Node 
        sheet['E4'] = ForceChoice[a][4] # Case
        
        
        workbook.save(filename = BoltCalcStr)
       
        # Will need to modify Loc String Below
        Loc = "C:/Users/brendanlia/Desktop/URGENT/"
        DFID = ForceStr + str(a) + ".xlsx"
        REFID = Loc + DFID
        workbook = load_workbook(filename = BoltCalcStr)
        workbook.save(filename = REFID)
        
        #ExcelRefresh
        time.sleep(0.2)
        refreshExcel1(REFID)
        
        # Excel to PDF
        PDF = ForceStr + str(a) + ".pdf"
        PDFID = Loc + PDF
        excel = client.Dispatch("Excel.Application")
        sheets = excel.Workbooks.Open(REFID)
        
        if sheetCalcStr == 'A-1':
            sheetindex = 1
        elif sheetCalcStr == 'A-2':
            sheetindex = 2
        elif sheetCalcStr == 'B':
            sheetindex = 3
        elif sheetCalcStr == 'C-1':
            sheetindex = 4
        elif sheetCalcStr == 'C-2':
            sheetindex = 5
        work_sheets = sheets.Worksheets[sheetindex]
        work_sheets.ExportAsFixedFormat(0, PDFID)
        sheets.Close(True)
        
        # Allow Excel to refresh
        time.sleep(0.2)
        refreshExcel()
        
        # Get UR Values and Calculate Max
        wb = load_workbook(filename = BoltCalcStr, data_only = True)
        ws = wb[sheetCalcStr]
        your_data = pandas.DataFrame(ws.values)
        TUR.append(your_data[6][52])
        BUR.append(your_data[6][67])
        SUR.append(your_data[6][68])
        print(a)
        
        
    # append to CombinedForces Excel
    workbook = load_workbook(filename = ForceCombinedStr)
    sheet = workbook[ForceStr]
    sheet["N1"] = "Max Bolt Tension UR"
    sheet["O1"] = "Max Plate Bending UR"
    sheet["P1"] = "Max Plate Shear UR"
    for a in range(len(TUR)):
        b = str(a+2)
        L = "N"
        M = "O"
        N = "P"
        d = L+b
        e = M+b
        f = N+b
        sheet[d] = TUR[a]
        sheet[e] = BUR[a]
        sheet[f] = SUR[a]
    
    workbook.save(filename = ForceCombinedStr)

'----------------------------------------------------------------------------------------------------------------------------------------------------'

# Get Min Max Forces Data

MinMaxLoc = r'C:\Users\brendanlia\Desktop\URGENT\CombinedMinMax.xlsx'

MaxFx = pandas.read_excel(MinMaxLoc, sheet_name = 'MaxFx')
MaxFy = pandas.read_excel(MinMaxLoc, sheet_name = 'MaxFy')
MaxFz = pandas.read_excel(MinMaxLoc, sheet_name = 'MaxFz')
MinFx = pandas.read_excel(MinMaxLoc, sheet_name = 'MinFx')
MinFy = pandas.read_excel(MinMaxLoc, sheet_name = 'MinFy')
MinFz = pandas.read_excel(MinMaxLoc, sheet_name = 'MinFz')
MaxMx = pandas.read_excel(MinMaxLoc, sheet_name = 'MaxMx')
MaxMy = pandas.read_excel(MinMaxLoc, sheet_name = 'MaxMy')
MinMx = pandas.read_excel(MinMaxLoc, sheet_name = 'MinMx')
MinMy = pandas.read_excel(MinMaxLoc, sheet_name = 'MinMy')


# Transpose for easier data access by column
MaxFx = pandas.DataFrame.transpose(MaxFx)
MaxFy = pandas.DataFrame.transpose(MaxFy)
MaxFz = pandas.DataFrame.transpose(MaxFz)
MinFx = pandas.DataFrame.transpose(MinFx)
MinFy = pandas.DataFrame.transpose(MinFy)
MinFz = pandas.DataFrame.transpose(MinFz)
MaxMx = pandas.DataFrame.transpose(MaxMx)
MaxMy = pandas.DataFrame.transpose(MaxMy)
MinMx = pandas.DataFrame.transpose(MinMx)
MinMy = pandas.DataFrame.transpose(MinMy)


getMaxUR(MaxFx, "BoltCalc3.xlsx", "CombinedMinMax.xlsx", "MaxFx")
print("Completed MaxFx")
getMaxUR(MaxFy, "BoltCalc3.xlsx", "CombinedMinMax.xlsx", "MaxFy")
print("Completed MaxFy")
getMaxUR(MaxFz, "BoltCalc3.xlsx", "CombinedMinMax.xlsx", "MaxFz")
print("Completed MaxFz")
getMaxUR(MinFx, "BoltCalc3.xlsx", "CombinedMinMax.xlsx", "MinFx")
print("Completed MinFx")
getMaxUR(MinFy, "BoltCalc3.xlsx", "CombinedMinMax.xlsx", "MinFy")
print("Completed MinFy")
getMaxUR(MinFz, "BoltCalc3.xlsx", "CombinedMinMax.xlsx", "MinFz")
print("Completed MinFz")
getMaxUR(MaxMx, "BoltCalc3.xlsx", "CombinedMinMax.xlsx", "MaxMx")
print("Completed MaxMx")
getMaxUR(MaxMy, "BoltCalc3.xlsx", "CombinedMinMax.xlsx", "MaxMy")
print("Completed MaxMy")
getMaxUR(MinMx, "BoltCalc3.xlsx", "CombinedMinMax.xlsx", "MinMx")
print("Completed MinMx")
getMaxUR(MinMy, "BoltCalc3.xlsx", "CombinedMinMax.xlsx", "MinMy")
print("Completed MinMy")


input('Press Enter to Close')
