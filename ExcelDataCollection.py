# -*- coding: utf-8 -*-
"""
Created on Wed Jul  7 09:17:36 2021

@author: brendanlia
"""

import numpy
import xlsxwriter
import pandas
import easygui 
import openpyxl
from openpyxl import load_workbook
import time
import win32com.client

'--------------------------------------------------------------------------------------------------------------------------------------------------'
#Def Functions

def refreshExcel():
    xlapp = win32com.client.DispatchEx("Excel.Application")
    wb = xlapp.Workbooks.Open(r"C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\BoltCalc1.xlsx")
    wb.RefreshAll()
    xlapp.CalculateUntilAsyncQueriesDone()
    wb.Save()
    xlapp.Quit()
    
'----------------------------------------------------------------------------------------------------------------------------------------------------'
    
def getMaxUR(ForceChoice, BoltCalcStr, ForceCombinedStr, ForceStr):
    # ForceChoice = selction of MaxFx, MaxFy, ... , MinFy
    # BoltCalcStr = file name of the bolt calc spreadsheet, input as a string e.g. "BoltCalc.xlsx"
    # SheetCalcStr = the name of the sheet within BoltCalcStr that is the desired bolt configuration, input as string, e.g. "BP (CAP14)"
    # ForceCombinedStr = the file name of the combined forces xlsx doc, input as string, e.g. "CombinedMinMax.xlsx"
    #ForceStr = the Force Choice in string format, input as string, e.g. "MaxFx"
    
    TUR = []
    BUR = []
    SUR = []
    
    for a in range(len(ForceChoice.columns)): #Input new Force Values
        Group = ForceChoice[a][10]
        size = len(Group)-2
        cut = Group[2:][:size - 2]
        sheetCalcStr = cut
        workbook = load_workbook(filename = BoltCalcStr)
        sheet = workbook[sheetCalcStr]
        if ForceChoice[a][6] > 0: #Compression
            sheet["F6"] = 0
            sheet["F7"] = abs(ForceChoice[a][6]) 
        elif ForceChoice[a][6] < 0: #Tension
            sheet["F6"] = abs(ForceChoice[a][6])
            sheet["F7"] = 0
        else:     
            sheet["F6"] = 0
            sheet["F7"] = 0
        
        sheet["F8"] = ForceChoice[a][4]
        sheet["F9"] = ForceChoice[a][5]
        sheet["F10"] = abs(ForceChoice[a][7])
        sheet["F11"] = abs(ForceChoice[a][8])
        workbook.save(filename = BoltCalcStr)
       
        #Allow Excel to refresh
        time.sleep(0.1)
        refreshExcel()
        
        #Get UR Values and Calculate Max
        wb = load_workbook(filename = BoltCalcStr, data_only = True)
        ws = wb[sheetCalcStr]
        your_data = pandas.DataFrame(ws.values)
        TUR.append(your_data[6][52])
        BUR.append(your_data[6][67])
        SUR.append(your_data[6][68])
        print(a)
        
        
    #append to CombinedForces Excel
    workbook = load_workbook(filename = ForceCombinedStr)
    sheet = workbook[ForceStr]
    sheet["L1"] = "Max Bolt Tension UR"
    sheet["M1"] = "Max Plate Bending UR"
    sheet["N1"] = "Max Plate Shear UR"
    for a in range(len(TUR)):
        b = str(a+2)
        L = "L"
        M = "M"
        N = "N"
        d = L+b
        e = M+b
        f = N+b
        sheet[d] = TUR[a]
        sheet[e] = BUR[a]
        sheet[f] = SUR[a]
    
    workbook.save(filename = ForceCombinedStr)

'----------------------------------------------------------------------------------------------------------------------------------------------------'
# Get Min Max Forces Data

MaxFx = pandas.read_excel(r'C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\CombinedMinMax.xlsx', sheet_name = 'MaxFx')
MaxFy = pandas.read_excel(r'C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\CombinedMinMax.xlsx', sheet_name = 'MaxFy')
MaxFz = pandas.read_excel(r'C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\CombinedMinMax.xlsx', sheet_name = 'MaxFz')
MinFx = pandas.read_excel(r'C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\CombinedMinMax.xlsx', sheet_name = 'MinFx')
MinFy = pandas.read_excel(r'C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\CombinedMinMax.xlsx', sheet_name = 'MinFy')
MinFz = pandas.read_excel(r'C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\CombinedMinMax.xlsx', sheet_name = 'MinFz')
MaxMx = pandas.read_excel(r'C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\CombinedMinMax.xlsx', sheet_name = 'MaxMx')
MaxMy = pandas.read_excel(r'C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\CombinedMinMax.xlsx', sheet_name = 'MaxMy')
MinMx = pandas.read_excel(r'C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\CombinedMinMax.xlsx', sheet_name = 'MinMx')
MinMy = pandas.read_excel(r'C:\Users\brendanlia\Desktop\Airport Excel Data\19 tree base result check\CombinedMinMax.xlsx', sheet_name = 'MinMy')

MaxFx = pandas.DataFrame.transpose(MaxFx)
MaxFy = pandas.DataFrame.transpose(MaxFy)
MaxFz = pandas.DataFrame.transpose(MaxFz)
MinFx = pandas.DataFrame.transpose(MinFx)
MinFy = pandas.DataFrame.transpose(MinFy)
MinFz = pandas.DataFrame.transpose(MinFz)
MaxMx = pandas.DataFrame.transpose(MaxMx)
MaxMy = pandas.DataFrame.transpose(MaxMy)
MinMx = pandas.DataFrame.transpose(MinMx)
MinMy = pandas.DataFrame.transpose(MinMy)

getMaxUR(MaxFx, "BoltCalc1.xlsx", "CombinedMinMax.xlsx", "MaxFx")
print("Complete MaxFx")
getMaxUR(MaxFy, "BoltCalc1.xlsx", "CombinedMinMax.xlsx", "MaxFy")
print("Complete MaxFy")
getMaxUR(MaxFz, "BoltCalc1.xlsx", "CombinedMinMax.xlsx", "MaxFz")
print("Complete MaxFz")
getMaxUR(MaxFx, "BoltCalc1.xlsx", "CombinedMinMax.xlsx", "MinFx")
print("Complete MinFx")
getMaxUR(MinFy, "BoltCalc1.xlsx", "CombinedMinMax.xlsx", "MinFy")
print("Complete MinFy")
getMaxUR(MinFz, "BoltCalc1.xlsx", "CombinedMinMax.xlsx", "MinFz")
print("Complete MinFz")
getMaxUR(MaxMx, "BoltCalc1.xlsx", "CombinedMinMax.xlsx", "MaxMx")
print("Complete MaxMx")
getMaxUR(MaxMy, "BoltCalc1.xlsx", "CombinedMinMax.xlsx", "MaxMy")
print("Complete MaxMy")
getMaxUR(MinMx, "BoltCalc1.xlsx", "CombinedMinMax.xlsx", "MinMx")
print("Complete MinMx")
getMaxUR(MinMy, "BoltCalc1.xlsx", "CombinedMinMax.xlsx", "MinMy")
print("Complete MinMy")

input('Press Enter to Close')
    









